import os
import json
import requests
import re
from flask import Flask, render_template, request, jsonify, send_file
from groq import Groq
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import threading

app = Flask(__name__)

# ─── CONFIG ───────────────────────────────────────────────
GROQ_API_KEY = os.getenv("GROQ_API_KEY")
SERPER_API_KEY = os.getenv("SERPER_API_KEY")
# ──────────────────────────────────────────────────────────

if not GROQ_API_KEY or not SERPER_API_KEY:
    raise RuntimeError("Set GROQ_API_KEY and SERPER_API_KEY environment variables before starting the app.")

client = Groq(api_key=GROQ_API_KEY)

# ── TOOL: Search the web ──────────────────────────────────
def search_web(query: str) -> str:
    url = "https://google.serper.dev/search"
    headers = {"X-API-KEY": SERPER_API_KEY, "Content-Type": "application/json"}
    payload = {"q": query, "num": 5}
    response = requests.post(url, headers=headers, json=payload)
    if response.status_code != 200:
        return "Search failed."
    data = response.json()
    results = []
    for item in data.get("organic", [])[:5]:
        results.append(f"Title: {item.get('title')}\nLink: {item.get('link')}\nSnippet: {item.get('snippet')}")
    return "\n\n".join(results) if results else "No results found."

# ── STAGE 1: Suggest companies ────────────────────────────
def suggest_companies(user_request: str) -> dict:
    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        max_tokens=1024,
        messages=[
            {
                "role": "system",
                "content": (
                    "You are a business research assistant. When the user describes what kind of companies they need "
                    "(suppliers, clients, distributors, etc.), suggest exactly 8 relevant Indian or global companies. "
                    "Respond ONLY with a JSON object like this:\n"
                    "{\n"
                    '  "title": "Short title describing the list",\n'
                    '  "reason": "One sentence why these companies are relevant",\n'
                    '  "companies": [\n'
                    '    {"name": "Company Name", "reason": "Why this company is relevant"},\n'
                    '    ...\n'
                    '  ]\n'
                    "}"
                )
            },
            {"role": "user", "content": user_request}
        ]
    )

    text = response.choices[0].message.content or ""
    try:
        json_match = re.search(r'\{.*\}', text, re.DOTALL)
        if json_match:
            return json.loads(json_match.group())
    except:
        pass
    return {"title": "Suggestions", "reason": "", "companies": []}

# ── STAGE 2: Research one company ─────────────────────────
def research_company(company_name: str) -> dict:
    tools = [
        {
            "type": "function",
            "function": {
                "name": "search_web",
                "description": "Search Google for information about a company.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "query": {"type": "string"}
                    },
                    "required": ["query"]
                }
            }
        }
    ]

    messages = [
        {
            "role": "system",
            "content": (
                "You are a lead research agent. Search for the company and extract: "
                "website, CEO or founder name, contact email, LinkedIn page, company size, industry. "
                "Respond ONLY with JSON:\n"
                "{\n"
                '  "company": "...", "website": "...", "ceo_name": "...",\n'
                '  "email": "...", "linkedin": "...", "industry": "...",\n'
                '  "company_size": "...", "summary": "One line about what they do"\n'
                "}"
            )
        },
        {"role": "user", "content": f"Research this company: {company_name}"}
    ]

    for _ in range(6):
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            max_tokens=1024,
            tools=tools,
            tool_choice="auto",
            messages=messages
        )

        msg = response.choices[0].message

        if msg.tool_calls:
            messages.append({
                "role": "assistant",
                "content": msg.content or "",
                "tool_calls": [
                    {
                        "id": tc.id,
                        "type": "function",
                        "function": {"name": tc.function.name, "arguments": tc.function.arguments}
                    } for tc in msg.tool_calls
                ]
            })
            for tool_call in msg.tool_calls:
                query  = json.loads(tool_call.function.arguments).get("query", "")
                result = search_web(query)
                messages.append({"role": "tool", "tool_call_id": tool_call.id, "content": result})
        else:
            text = msg.content or ""
            try:
                json_match = re.search(r'\{.*\}', text, re.DOTALL)
                if json_match:
                    return json.loads(json_match.group())
            except:
                pass
            break

    return {
        "company": company_name, "website": "Not found", "ceo_name": "Not found",
        "email": "Not found", "linkedin": "Not found", "industry": "Not found",
        "company_size": "Not found", "summary": "Could not extract details"
    }

# ── EXCEL EXPORT ──────────────────────────────────────────
def export_to_excel(leads: list) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    headers = ["Company", "Website", "CEO / Founder", "Email", "LinkedIn", "Industry", "Company Size", "Summary"]
    header_fill = PatternFill("solid", fgColor="1a1a2e")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row, lead in enumerate(leads, 2):
        ws.cell(row=row, column=1, value=lead.get("company", ""))
        ws.cell(row=row, column=2, value=lead.get("website", ""))
        ws.cell(row=row, column=3, value=lead.get("ceo_name", ""))
        ws.cell(row=row, column=4, value=lead.get("email", ""))
        ws.cell(row=row, column=5, value=lead.get("linkedin", ""))
        ws.cell(row=row, column=6, value=lead.get("industry", ""))
        ws.cell(row=row, column=7, value=lead.get("company_size", ""))
        ws.cell(row=row, column=8, value=lead.get("summary", ""))
        if row % 2 == 0:
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="f0f0f0")
    widths = [25, 30, 20, 30, 35, 20, 18, 50]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    filename = f"leads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join("outputs", filename)
    os.makedirs("outputs", exist_ok=True)
    wb.save(filepath)
    return filepath

# ── FLASK ROUTES ──────────────────────────────────────────
results_store = {}

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/suggest", methods=["POST"])
def suggest():
    data    = request.json
    prompt  = data.get("prompt", "").strip()
    if not prompt:
        return jsonify({"error": "No prompt provided"}), 400
    result = suggest_companies(prompt)
    return jsonify(result)

@app.route("/research", methods=["POST"])
def research():
    data      = request.json
    companies = [c.strip() for c in data.get("companies", []) if c.strip()]
    if not companies:
        return jsonify({"error": "No companies provided"}), 400
    session_id = datetime.now().strftime("%Y%m%d%H%M%S")
    results_store[session_id] = {"status": "running", "results": [], "total": len(companies), "current": ""}

    def run_research():
        results = []
        for i, company in enumerate(companies):
            results_store[session_id]["current"] = f"Researching: {company} ({i+1}/{len(companies)})"
            lead = research_company(company)
            results.append(lead)
            results_store[session_id]["results"] = results
        filepath = export_to_excel(results)
        results_store[session_id].update({"status": "done", "file": filepath, "current": "Done!"})

    threading.Thread(target=run_research).start()
    return jsonify({"session_id": session_id})

@app.route("/status/<session_id>")
def status(session_id):
    return jsonify(results_store.get(session_id, {"status": "not_found"}))

@app.route("/download/<session_id>")
def download(session_id):
    session = results_store.get(session_id)
    if session and session.get("file"):
        return send_file(session["file"], as_attachment=True)
    return "File not found", 404

if __name__ == "__main__":
    app.run(debug=True, port=5000)
